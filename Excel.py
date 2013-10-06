from __future__ import print_function

import xlrd
import xml.etree.ElementTree
from openpyxl.reader.excel import load_workbook
import os
import math
import itertools
import unicodedata

def toAscii( s ):
	if s is None or s == '':
		return ''
	ret = unicodedata.normalize('NFKD', s).encode('ascii','ignore') if type(s) == unicode else str(s)
	if ret.endswith( '.0' ):
		ret = ret[:-2]
	return ret

#----------------------------------------------------------------------------

class ReadExcelBase( object ):
	def __init__( self, filename ):
		pass
		
	def sheet_names( self ):
		assert( False )

	def iter_list( self ):
		assert( False )
		
#----------------------------------------------------------------------------

class ReadExcelXls( ReadExcelBase ):
	def __init__(self, filename):
		ReadExcelBase.__init__( self, filename )
		if not os.path.isfile(filename):
			raise ValueError, "%s is not a valid filename" % filename
		self.book = xlrd.open_workbook(filename)
		
	def is_nonempty_row(self, sheet, i):
		values = sheet.row_values(i)
		if isinstance(values[0], basestring) and values[0].startswith('#'):
			return False # ignorable comment row
		return any( bool(v) for v in values )
	
	def sheet_names( self ):
		return self.book.sheet_names()
		
	def _parse_row(self, sheet, row_index, date_as_tuple):
		""" Sanitize incoming excel data """
		# Data Type Codes:
		#  EMPTY 0
		#  TEXT 1 a Unicode string
		#  NUMBER 2 float
		#  DATE 3 float
		#  BOOLEAN 4 int; 1 means TRUE, 0 means FALSE
		#  ERROR 5
		values = []
		for type, value in itertools.izip(sheet.row_types(row_index), sheet.row_values(row_index)):
			if type == 2:
				if value == int(value):
					value = int(value)
			elif type == 3:
				if isinstance(value, float) and value < 1.0:
					t = value * (24.0*60.0*60.0)
					if int(t + 0.000001) == int(t+1.0):
						secs = int(t + 0.000001)
						fract = 0.0
					else:
						fract, secs = math.modf( t )
						if fract < 0.000000001:
							fract = 0.0
						secs = int(secs)
					if fract:
						value = '%02d:%02d:%02d.%s' % ( secs // (60*60), (secs // 60) % 60, secs % 60, ('%.20f'%fract)[2:])
					else:
						value = '%02d:%02d:%02d' % (secs // (60*60), (secs // 60) % 60, secs % 60)
				else:
					try:
						datetuple = xlrd.xldate_as_tuple(value, self.book.datemode)
						validDate = True
					except:
						value = 'UnreadableDate'
						validDate = False
					if validDate:
						if date_as_tuple:
							value = datetuple
						else:
							# time only - no date component
							if datetuple[0] == 0 and datetuple[1] == 0 and  datetuple[2] == 0:
								value = "%02d:%02d:%02d" % datetuple[3:]
							# date only, no time
							elif datetuple[3] == 0 and datetuple[4] == 0 and datetuple[5] == 0:
								value = "%04d/%02d/%02d" % datetuple[:3]
							else: # full date
								value = "%04d/%02d/%02d %02d:%02d:%02d" % datetuple
			elif type == 5:
				value = xlrd.error_text_from_code[value]
			values.append(value)
		return values

	def iter_list(self, sname, date_as_tuple=False):
		sheet = self.book.sheet_by_name(sname) # XLRDError
		for i in range(sheet.nrows):
			yield self._parse_row(sheet, i, date_as_tuple)

#----------------------------------------------------------------------------

class ReadExcelXlsx( ReadExcelBase ):
	def __init__(self, filename):
		ReadExcelBase.__init__( self, filename )
		if not os.path.isfile(filename):
			raise ValueError, "%s is not a valid filename" % filename
		self.book = load_workbook( filename = filename, use_iterators = True )
		
	def sheet_names( self ):
		return self.book.get_sheet_names()
	
	def iter_list(self, sname):
		ws = self.book.get_sheet_by_name( name = sname )
		for row in ws.iter_rows():
			r = [cell.internal_value for cell in row]
			yield r

#----------------------------------------------------------------------------

def GetExcelReader( filename ):
	if filename.endswith( '.xls' ):
		return ReadExcelXls( filename )
	elif filename.endswith( '.xlsx' ) or filename.endswith( '.xlsm' ):
		return ReadExcelXlsx( filename )
	else:
		raise ValueError, '%s is not a recognized Excel format' % filename
		
#----------------------------------------------------------------------------

